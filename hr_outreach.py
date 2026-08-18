import argparse, os, re, random, time, json
from pathlib import Path
from datetime import datetime
from email.message import EmailMessage
import smtplib

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE = Path(__file__).resolve().parent
load_dotenv(BASE / ".env")

INPUT_XLSX = BASE / "hr details.xlsx"
TRACKER_XLSX = BASE / os.getenv("TRACKER_PATH", "outreach_tracker.xlsx")
CV_PATH = BASE / os.getenv("CV_PATH", "SAURAV KUMAR UPDATED CV.pdf")

SENDER_NAME = os.getenv("SENDER_NAME", "Saurav Kumar")
SENDER_EMAIL = os.getenv("SENDER_EMAIL", "")
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", SENDER_EMAIL)
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_DELAY_MIN = int(os.getenv("EMAIL_DELAY_MIN", "10"))
EMAIL_DELAY_MAX = int(os.getenv("EMAIL_DELAY_MAX", "15"))
WA_DELAY_MIN = int(os.getenv("WHATSAPP_DELAY_MIN", "15"))
WA_DELAY_MAX = int(os.getenv("WHATSAPP_DELAY_MAX", "20"))
TIMEOUT = int(os.getenv("WHATSAPP_TIMEOUT_MS", "120000"))
PROFILE = BASE / "whatsapp_profile"

# Debug Configuration
DEBUG_ENABLED = os.getenv("DEBUG_ENABLED", "false").lower() == "true"
DEBUG_OUTPUT_DIR = BASE / os.getenv("DEBUG_OUTPUT_DIR", "debug_logs")

SUBJECT = os.getenv("EMAIL_SUBJECT", "Application for Cloud / DevOps Engineer Opportunities – Saurav Kumar")

HEADERS = ["SL No.","HR Name","Contact No","Email Id","Company Name","Location","Remark",
           "Personalized Email Subject","Personalized Email","Personalized WhatsApp",
           "Email Status","Email Sent At","Email Error","WhatsApp Status","WhatsApp Sent At","WhatsApp Error"]

def st(v): return "" if v is None else str(v).strip()
def phone(v):
    d = re.sub(r"\D","",st(v))
    return ("91"+d) if len(d)==10 and d[0] in "6789" else d

def email_body(company):
    company = st(company) or "your organization"
    return f"""Dear Hiring Team,

I hope you are doing well.

I am Saurav Kumar, a Cloud / DevOps professional with experience in cloud operations, infrastructure support, monitoring, CI/CD, database support, troubleshooting, and production operations.

I am currently exploring suitable Cloud / DevOps opportunities and would be grateful if you could consider my profile for relevant openings at {company}.

Please find my CV attached for your reference. I would appreciate the opportunity to discuss any suitable position matching my skills and experience.

Thank you for your time and consideration.

Regards,
Saurav Kumar
{SENDER_EMAIL}"""

def wa_body(company):
    company = st(company) or "your organization"
    return f"""Hello,

I'm Saurav Kumar, a Cloud / DevOps professional. I'm currently exploring suitable Cloud / DevOps opportunities and wanted to share my profile for consideration at {company}.

I've attached my CV for your reference. If there is any relevant opening matching my experience, I would be grateful for an opportunity to discuss it.

Thank you.

Regards,
Saurav Kumar"""

def prepare():
    if TRACKER_XLSX.exists():
        return
    src = load_workbook(INPUT_XLSX, read_only=True).active
    rows = list(src.iter_rows(values_only=True))
    hdr = {st(v):i for i,v in enumerate(rows[0])}
    records=[]
    for n,row in enumerate(rows[1:],1):
        if not any(st(x) for x in row): continue
        def get(k): return row[hdr[k]] if k in hdr and hdr[k] < len(row) else ""
        company=st(get("Company Name"))
        records.append([st(get("SL No.")) or str(n),st(get("HR Name")),st(get("Contact No")),
                        st(get("Email Id")),company,st(get("Location")),st(get("Remark")),
                        SUBJECT,email_body(company),wa_body(company),"Pending","","","Pending","",""])
    wb=load_workbook(INPUT_XLSX)
    ws=wb.active
    ws.delete_rows(1,ws.max_row)
    for c,h in enumerate(HEADERS,1): ws.cell(1,c).value=h
    for r,row in enumerate(records,2):
        for c,v in enumerate(row,1): ws.cell(r,c).value=v
    wb.save(TRACKER_XLSX)
    print(f"Prepared {len(records)} HR records.")

def get_tracker():
    if not TRACKER_XLSX.exists(): prepare()
    wb=load_workbook(TRACKER_XLSX)
    ws=wb.active
    cols={st(ws.cell(1,c).value):c for c in range(1,ws.max_column+1)}
    recs=[]
    for r in range(2,ws.max_row+1):
        rec={h:ws.cell(r,c).value for h,c in cols.items()}; rec["_row"]=r; recs.append(rec)
    return wb,ws,recs,cols

def write_status(wb,ws,cols,row,status_col,time_col,error_col,val,error=""):
    ws.cell(row,cols[status_col]).value=val
    ws.cell(row,cols[time_col]).value=datetime.now().strftime("%Y-%m-%d %H:%M:%S") if val=="Sent" else ""
    ws.cell(row,cols[error_col]).value=st(error)[:1500]
    wb.save(TRACKER_XLSX)

def send_email(to,subject,body):
    msg=EmailMessage()
    msg["From"]=f"{SENDER_NAME} <{SENDER_EMAIL}>"; msg["To"]=to; msg["Subject"]=subject
    msg.set_content(body)
    with open(CV_PATH,"rb") as f:
        msg.add_attachment(f.read(),maintype="application",subtype="pdf",filename=CV_PATH.name)
    with smtplib.SMTP(SMTP_HOST,SMTP_PORT,timeout=60) as smtp:
        smtp.starttls(); smtp.login(SMTP_USERNAME,SMTP_PASSWORD); smtp.send_message(msg)

def debug(page,number,tag):
    if not DEBUG_ENABLED:
        return
    
    DEBUG_OUTPUT_DIR.mkdir(exist_ok=True)
    safe=re.sub(r"\D","",number) or "unknown"
    
    try: 
        page.screenshot(path=str(DEBUG_OUTPUT_DIR/f"whatsapp_debug_{safe}_{tag}.png"))
    except: 
        pass
    
    try:
        (DEBUG_OUTPUT_DIR/f"whatsapp_debug_{safe}_{tag}.html").write_text(page.content(),encoding="utf-8")
    except: 
        pass

# WhatsApp Web renders two different Send buttons: one inside <footer> for the
# typed message, and one inside the attachment preview overlay. They must never
# be confused, so every lookup below is explicitly scoped to one or the other.
SEND_SELECTORS=['div[role="button"][aria-label="Send"]','button[aria-label="Send"]',
                '[role="button"][aria-label*="Send" i]','button[aria-label*="Send" i]',
                'button[title*="Send" i]','[data-testid="send"]',
                'span[data-icon="send"]','span[data-icon="send-filled"]',
                'span[data-icon="wds-ic-send-filled"]','span[data-icon*="send" i]']

CAPTION_SELECTORS=['div[contenteditable="true"][aria-label*="caption" i]',
                   'div[contenteditable="true"][role="textbox"]']

ATTACH_SELECTORS=['button[aria-label*="Attach" i]','div[role="button"][aria-label*="Attach" i]',
                  'button[title*="Attach" i]','[data-testid="clip"]',
                  'span[data-icon="clip"]','span[data-icon="plus"]','span[data-icon="plus-rounded"]']

DOC_MENU_SELECTORS=['li[data-testid="mi-attach-document"]','[aria-label*="Document" i]',
                    'div[role="button"]:has-text("Document")','li:has-text("Document")']

def first_visible(page,selectors):
    for sel in selectors:
        try:
            loc=page.locator(sel)
            for i in range(loc.count()):
                if loc.nth(i).is_visible(): return loc.nth(i)
        except: pass
    return None

def clickable(el):
    # Icon <span>s are not the click target; resolve to the button around them.
    try:
        if el.evaluate("e => e.tagName.toLowerCase()") in ("span","svg","path"):
            anc=el.locator('xpath=ancestor-or-self::*[self::button or @role="button"][1]')
            if anc.count() and anc.first.is_visible(): return anc.first
    except: pass
    return el

def in_footer(el):
    try: return bool(el.evaluate("e => !!e.closest('footer')"))
    except: return False

def find_send(page,composer):
    # composer=True  -> the Send inside the attachment preview
    # composer=False -> the Send inside the chat footer
    for sel in SEND_SELECTORS:
        try: loc=page.locator(sel)
        except: continue
        for i in range(loc.count()):
            try:
                el=loc.nth(i)
                if not el.is_visible(): continue
                btn=clickable(el)
                if in_footer(btn)==composer: continue
                return btn
            except: pass
    return None

def caption_box(page):
    # Only a text box OUTSIDE <footer> belongs to the attachment preview. The
    # footer's own input is always visible, so counting it here would make
    # composer_open() permanently true and every send look like a failure.
    for sel in CAPTION_SELECTORS:
        try:
            loc=page.locator(sel)
            for i in range(loc.count()):
                el=loc.nth(i)
                if el.is_visible() and not in_footer(el): return el
        except: pass
    return None

def composer_open(page):
    return find_send(page,True) is not None or caption_box(page) is not None

def outgoing_count(page):
    # Outgoing rows carry class "message-out" / data-id starting with "true_".
    for sel in ("div.message-out",'div[class*="message-out"]','div[data-id^="true_"]'):
        try:
            n=page.locator(sel).count()
            if n: return n
        except: pass
    return 0

def footer_box(page):
    loc=page.locator('footer div[contenteditable="true"]')
    try:
        for i in range(loc.count()-1,-1,-1):
            if loc.nth(i).is_visible(): return loc.nth(i)
    except: pass
    return None

def box_text(page):
    b=footer_box(page)
    if b is None: return ""
    try: return (b.inner_text() or "").strip()
    except: return ""

def invalid_number_reason(page):
    try: txt=page.locator('div[role="dialog"]').first.inner_text(timeout=1500)
    except: return ""
    if re.search(r"invalid|not on whatsapp|isn't on whatsapp|no result",txt,re.I):
        return txt.strip().splitlines()[0]
    return ""

def wait_chat(page):
    end=time.time()+90
    while time.time()<end:
        if invalid_number_reason(page): return None
        for sel in ['footer div[contenteditable="true"]','div[contenteditable="true"][role="textbox"]']:
            try:
                loc=page.locator(sel)
                for i in range(loc.count()-1,-1,-1):
                    if loc.nth(i).is_visible(): return loc.nth(i)
            except: pass
        page.wait_for_timeout(500)
    return None

def type_message(page,box,message):
    # insert_text keeps unicode intact and Shift+Enter adds a newline without
    # sending, so a multi-line message never goes out as several messages.
    box.click()
    try: box.fill("")
    except: pass
    for i,line in enumerate(message.replace("\r\n","\n").split("\n")):
        if i: page.keyboard.press("Shift+Enter")
        if line: page.keyboard.insert_text(line)

def send_text(page,box,message,number):
    before=outgoing_count(page)
    type_message(page,box,message)
    page.wait_for_timeout(400)
    if not box_text(page):
        debug(page,number,"type_failed")
        raise RuntimeError("Could not type the WhatsApp message into the chat box.")
    btn=find_send(page,False)
    try:
        if btn: btn.click(timeout=5000)
        else: page.keyboard.press("Enter")
    except: page.keyboard.press("Enter")
    end=time.time()+20
    while time.time()<end:
        if outgoing_count(page)>before or not box_text(page): return
        page.wait_for_timeout(400)
    debug(page,number,"text_failed")
    raise RuntimeError("WhatsApp text message could not be confirmed as sent.")

def attach_pdf(page,number):
    btn=first_visible(page,ATTACH_SELECTORS)
    if not btn:
        debug(page,number,"attach_failed")
        raise RuntimeError("WhatsApp Attach control not found")
    clickable(btn).click()
    page.wait_for_timeout(800)

    # Preferred path: let WhatsApp open its own chooser for the Document item.
    doc=first_visible(page,DOC_MENU_SELECTORS)
    if doc:
        try:
            with page.expect_file_chooser(timeout=8000) as fc:
                clickable(doc).click()
            fc.value.set_files(str(CV_PATH))
            return
        except: pass

    # Fallback: pick the input that accepts any file type. The photos/videos
    # input restricts accept to image/video and silently drops a PDF, which is
    # why no preview - and therefore no Send button - ever appears.
    inputs=page.locator('input[type="file"]')
    ranked=[]
    for i in range(inputs.count()):
        try: acc=(inputs.nth(i).get_attribute("accept") or "").lower()
        except: continue
        if "image/" in acc or "video/" in acc: score=2
        elif acc in ("","*","*/*") or "pdf" in acc or "application/" in acc: score=0
        else: score=1
        ranked.append((score,i))
    for _,i in sorted(ranked):
        try:
            inputs.nth(i).set_input_files(str(CV_PATH))
            page.wait_for_timeout(1200)
            if composer_open(page): return
        except: pass
    debug(page,number,"attach_failed")
    raise RuntimeError("WhatsApp document input not found, or it rejected the CV.")

def document_sent(page,before):
    # A real confirmation needs both: the preview closed AND a new outgoing row.
    end=time.time()+15
    while time.time()<end:
        if not composer_open(page) and outgoing_count(page)>before: return True
        page.wait_for_timeout(400)
    return False

def send_document(page,number):
    # The composer's Send button only exists once WhatsApp has built the
    # preview, so wait for the composer itself rather than for any Send button.
    end=time.time()+45
    while time.time()<end:
        if composer_open(page): break
        page.wait_for_timeout(500)
    else:
        debug(page,number,"preview_failed")
        raise RuntimeError("WhatsApp attachment preview never opened.")

    before=outgoing_count(page)
    deadline=time.time()+45
    while time.time()<deadline:
        btn=find_send(page,True)
        if btn:
            try: btn.scroll_into_view_if_needed(timeout=2000)
            except: pass
            for attempt in (lambda: btn.click(timeout=5000),
                            lambda: btn.click(timeout=5000,force=True),
                            lambda: btn.press("Enter")):
                try: attempt()
                except: continue
                if document_sent(page,before): return
        page.wait_for_timeout(700)

    # Keyboard fallback: focus the caption box, then press Enter.
    try:
        cap=caption_box(page)
        if cap: cap.click()
        page.keyboard.press("Enter")
        if document_sent(page,before): return
    except: pass

    debug(page,number,"send_failed")
    raise RuntimeError("CV was attached but Send could not be confirmed. Debug PNG/HTML saved.")

def send_whatsapp(page,number,message):
    page.goto(f"https://web.whatsapp.com/send?phone={number}",wait_until="domcontentloaded",timeout=TIMEOUT)
    box=wait_chat(page)
    if not box:
        debug(page,number,"chat_failed")
        raise RuntimeError(invalid_number_reason(page) or "WhatsApp chat did not load.")
    # Text goes out on its own first. Typing it before attaching leaves a live
    # footer Send button behind the preview overlay, and that stale button is
    # what the automation used to click instead of the attachment's Send.
    send_text(page,box,message,number)
    page.wait_for_timeout(800)
    attach_pdf(page,number)
    send_document(page,number)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--mode",choices=["prepare","both"],default="prepare")
    ap.add_argument("--limit",type=int,default=1)
    args=ap.parse_args()
    if args.mode=="prepare": prepare(); return

    if not CV_PATH.exists(): raise SystemExit(f"CV not found: {CV_PATH}")
    wb,ws,recs,cols=get_tracker()
    with sync_playwright() as p:
        context=p.chromium.launch_persistent_context(
            str(PROFILE),headless=False,viewport=None,
            args=["--start-maximized","--disable-notifications"]
        )
        page=context.pages[0] if context.pages else context.new_page()
        page.goto("https://web.whatsapp.com",wait_until="domcontentloaded",timeout=TIMEOUT)
        print("WhatsApp Web opened. Scan QR only if required.")
        input("Press Enter when WhatsApp Web is ready: ")

        processed=0
        for rec in recs:
            if processed>=args.limit: break
            if st(rec.get("Email Status"))=="Sent" and st(rec.get("WhatsApp Status"))=="Sent": continue
            print(f"\n========== HR {rec.get('SL No.','')} | {rec.get('Company Name','')} ==========")

            if st(rec.get("Email Status"))!="Sent":
                to=st(rec.get("Email Id"))
                if to and "@" in to:
                    try:
                        send_email(to,st(rec.get("Personalized Email Subject")) or SUBJECT,st(rec.get("Personalized Email")))
                        write_status(wb,ws,cols,rec["_row"],"Email Status","Email Sent At","Email Error","Sent")
                        print("Email: SENT ->",to)
                        time.sleep(random.randint(EMAIL_DELAY_MIN,EMAIL_DELAY_MAX))
                    except Exception as e:
                        write_status(wb,ws,cols,rec["_row"],"Email Status","Email Sent At","Email Error","Failed",str(e))
                        print("Email: FAILED ->",e)
                else:
                    write_status(wb,ws,cols,rec["_row"],"Email Status","Email Sent At","Email Error","Skipped","Invalid/missing email")

            if st(rec.get("WhatsApp Status"))!="Sent":
                number=phone(rec.get("Contact No"))
                if number:
                    try:
                        print("WhatsApp: preparing +"+number)
                        send_whatsapp(page,number,st(rec.get("Personalized WhatsApp")))
                        write_status(wb,ws,cols,rec["_row"],"WhatsApp Status","WhatsApp Sent At","WhatsApp Error","Sent")
                        print("WhatsApp: SENT -> +"+number)
                        time.sleep(random.randint(WA_DELAY_MIN,WA_DELAY_MAX))
                    except Exception as e:
                        write_status(wb,ws,cols,rec["_row"],"WhatsApp Status","WhatsApp Sent At","WhatsApp Error","Failed",str(e))
                        print("WhatsApp: FAILED ->",e)
                else:
                    write_status(wb,ws,cols,rec["_row"],"WhatsApp Status","WhatsApp Sent At","WhatsApp Error","Skipped","Invalid/missing phone")
            processed+=1
        context.close()
    print("Both-channel run complete.")

if __name__=="__main__": main()
